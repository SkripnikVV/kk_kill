import datetime, re, glob, os, sys, json, time
from difflib import SequenceMatcher as SM
from multiprocessing import Pool, cpu_count, Process, Queue, current_process
import sqlite3

import openpyxl as opx
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Font, Border, Side
import colorama
from colorama import Fore, Back, Style

root = r'X:\!Day\ALL_TOPIC\*'
excelpath = r'x:\Проверки'
spammer = r'\\192.168.0.230\archi\alcu\alcuSendArchiMail.exe'
workers = [['NSK-ВВОД', ['АЛЕКСЕЕВА']], ['NSK-ПОДКЛЮЧЕНИЕ', ['ЛАПИЦКАЯ', 'ЮЖИНА']], ['KHA-ВВОД', ['ГНИДЮК']],
           ['KHA-ПОДКЛЮЧЕНИЕ', ['МЕДВЕДЕВА']],
           ['TUVA-ВВОД', ['ГНИДЮК']], ['TUVA-ПОДКЛЮЧЕНИЕ', ['МЕДВЕДЕВА']], ['UDM-ВВОД', ['САРКИСЯН']],
           ['UDM-ПОДКЛЮЧЕНИЕ', ['НАУМОВА', 'СТАРИЦЫНА']],
           ['HBK-ВВОД', ['ЛОБОДА']], ['HBK-ПОДКЛЮЧЕНИЕ', ['ОГНЯННИКОВА', 'ТАТАРЧЕНКО']],
           ['MRK-ВВОД', ['ШИХАЛЕВА', 'КОЗЫРЕВА']],
           ['MRK-ПОДКЛЮЧЕНИЕ', ['ГАЛКИНА']], ['KMK-ВВОД', ['КОЛМАКОВ']], ['KMK-ПОДКЛЮЧЕНИЕ', ['ДУБИНСКАЯ']]]

eta_main = {
    'NSK': {
        "belongs": ['NSK', 'NSK_NAL'],
        "srcparent": "ОРГАНЫ ВЛАСТИ НОВОСИБИРСКОЙ ОБЛАСТИ",
        "publparent": "ИЗДАНИЯ НОВОСИБИРСКОЙ ОБЛАСТИ",
        "dia": [[7000000, 7199999], [47400000, 47499999], [47500000, 47549999], [74458860, 74448861]]
    },
    'HBK': {
        "belongs": ['HBK'],
        "srcparent": "ОРГАНЫ ВЛАСТИ ХАБАРОВСКОГО КРАЯ",
        "publparent": "ИЗДАНИЯ ХАБАРОВСКОГО КРАЯ",
        "dia": [[25400000, 25599999], [47600000, 47699999], [74725629, 74731628]]
    },
    'KHA': {
        "belongs": ['KHA'],
        "srcparent": "ОРГАНЫ ВЛАСТИ РЕСПУБЛИКИ ХАКАСИЯ",
        "publparent": "ИЗДАНИЯ РЕСПУБЛИКИ ХАКАСИЯ",
        "dia": [[20400000, 20599999], [47750000, 47799999]]
    },
    'KMK': {
        "belongs": ['KMK'],
        "srcparent": "ОРГАНЫ ВЛАСТИ РЕСПУБЛИКИ КАЛМЫКИЯ",
        "publparent": "ИЗДАНИЯ РЕСПУБЛИКИ КАЛМЫКИЯ",
        "dia": [[24800000, 24999999], [47550000, 47599999]]
    },
    'MRK': {
        "belongs": ['MRK'],
        "srcparent": "ОРГАНЫ ВЛАСТИ МУРМАНСКОЙ ОБЛАСТИ",
        "publparent": "ИЗДАНИЯ МУРМАНСКОЙ ОБЛАСТИ",
        "dia": [[16800000, 16999999], [44550000, 44599999]]
    },
    'TUVA': {
        "belongs": ['TUVA'],
        "srcparent": "ОРГАНЫ ВЛАСТИ РЕСПУБЛИКИ ТЫВА",
        "publparent": "ИЗДАНИЯ РЕСПУБЛИКИ ТЫВА",
        "dia": [[28600000, 28799999], [47700000, 47749999]]
    },
    'UDM': {
        "belongs": ['UDM', 'UDM_NAL', 'UDM_ARH'],
        "srcparent": "ОРГАНЫ ВЛАСТИ УДМУРТСКОЙ РЕСПУБЛИКИ",
        "publparent": "ИЗДАНИЯ УДМУРТСКОЙ РЕСПУБЛИКИ",
        "dia": [[15600000, 15799999], [47800000, 47899999]]
    }
}

font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
            color='FF000000')
fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
border = Border(left=Side(border_style=None, color='FF000000'),
                right=Side(border_style=None, color='FF000000'),
                top=Side(border_style=None, color='FF000000'),
                bottom=Side(border_style=None, color='FF000000'),
                diagonal=Side(border_style=None, color='FF000000'),
                diagonal_direction=0, outline=Side(border_style=None, color='FF000000'),
                vertical=Side(border_style=None, color='FF000000'),
                horizontal=Side(border_style=None, color='FF000000'))
alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False,
                      indent=0)
number_format = 'General'
protection = Protection(locked=True, hidden=False)
tablehead = NamedStyle(name="tablehead")
tablehead.font = Font(bold=True, size=14)
bd = Side(style='thick', color="000000")
tablehead.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True,
                                shrink_to_fit=False, indent=0)
tablehead.border = Border(left=bd, top=bd, right=bd, bottom=bd)
tablehead.fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')

tablecolhead = NamedStyle(name="tablecolhead")
tablecolhead.font = tablehead.font
tablecolhead.border = tablehead.border
tablecolhead.fill = PatternFill(start_color='F0E68C', end_color='F0E68C', fill_type='solid')


def title(path, message):
    pname = os.path.split(path)[1]
    print(Fore.LIGHTYELLOW_EX, Back.BLUE, '\u250c' + '\u2500' * (len(pname) + len(message) + 3) + '\u2510',
          Style.RESET_ALL)
    print(Fore.LIGHTYELLOW_EX, Back.BLUE, '\u2502 ' + pname + ' ' + Fore.LIGHTWHITE_EX + message + ' \u2502',
          Style.RESET_ALL)
    print(Fore.LIGHTYELLOW_EX, Back.BLUE, '\u2514' + '\u2500' * (len(pname) + len(message) + 3) + '\u2518',
          Style.RESET_ALL)


def printparam(param, value):
    print(Fore.WHITE, param, Fore.LIGHTWHITE_EX, value, Style.RESET_ALL)


def printerror(param, value):
    print(Fore.RED, param, Fore.LIGHTWHITE_EX, value, Style.RESET_ALL)

# def pipeout(topic, target,)


def killtr(deletetopics, savedeleted, filesqueue):
    # print(current_process().name,' started')
    deldict = {}
    for topic in deletetopics:
        deldict[topic] = 0
    # print(current_process().name,f' должно быть удалено: {len(deldict)} топиков.')
    while True:
        if filesqueue.qsize() == 0:
            break
        file = filesqueue.get()
        # print(filesqueue.qsize(),file)
        with open(file, 'r', encoding='cp866', errors='ignore') as fl:
            lines = fl.readlines()
        buff = []
        result = []
        skip = False
        delcount = 0
        # while True:
        for line in lines:
            # line=lines.pop(0)
            if line != '' and line[0] == '!':
                rsm = re.search('^!(TOPIC|OBJTOPIC|ANNOTOPIC)+\s+(\d+)', line)
                if rsm:
                    if rsm[2] in deldict.keys():
                        skip = True
                        delcount += 1
                        print(f'{rsm[0]}\tdeleted from: {file}')
                    else:
                        skip = False
            if skip == True:
                if savedeleted == True:
                    buff.append(line)
            else:
                result.append(line)
        if delcount > 0:
            with open(file, 'w', encoding='cp866', errors='ignore') as fl:
                fl.writelines(result)
            if savedeleted == True:
                with open(file + '.del', 'w', encoding='cp866', errors='ignore') as fl:
                    fl.writelines(buff)

    print(current_process().name, ' finished')


def sendmail2archi(path):
    os.system(
        spammer + f' -srv:192.168.0.230 -port:32100 -l:PIPEOUT -p:PIPEOUT -r:tron -s:"Отправление ошибок дневной дельты конечным пользователям" -b:"{path}"')
    os.system(
        spammer + f' -srv:192.168.0.230 -port:32100 -l:PIPEOUT -p:PIPEOUT -r:Compiler -s:"Отправление ошибок дневной дельты конечным пользователям" -b:"{path}"')
    os.system(
        spammer + f' -srv:192.168.0.230 -port:32100 -l:PIPEOUT -p:PIPEOUT -r:СУДАКОВА -s:"Отправление ошибок дневной дельты конечным пользователям" -b:"{path}"')
    os.system(
        spammer + f' -srv:192.168.0.230 -port:32100 -l:PIPEOUT -p:PIPEOUT -r:МЕДВЕДЬ -s:"Отправление ошибок дневной дельты конечным пользователям" -b:"{path}"')
    for worker in workers:
        mask, users = worker
        print(mask)
        print(worker)
        files = glob.glob(f'{path}\\*{mask}*.xlsx')
        print(files)
        if len(files) > 0:
            for user in users:
                with open(f'{path}\\mail.txt', 'w', encoding='cp1251') as fl:
                    fl.write('\nДорогой Друг!\n\nПосмотри, пожалуйста, эти файлы, что я приготовил для Тебя.\n\n')
                    fl.write('\n'.join(files))
                    fl.write('\n\nИсправлять ошибки нужно вдумчиво, чтобы не наделать новых.\n')
                    fl.write('Удачи тебе в этом нелёгком деле!\n')
                os.system(
                    spammer + f' -srv:192.168.0.230 -port:32100 -l:PIPEOUT -p:PIPEOUT -r:{user} -s:"Ошибки дневной дельты для Вас, {user}" -bf:"{path}\\mail.txt"')
    print('allex')


def generatekillcmt(data):
    # {'topic': topic, 'sub': sub, 'killdate': killdate, 'killname': killname}
    cmt = data['killname'].split()
    cmt[0] = chr(4) + cmt[0] + chr(4) + data['killtopic'] + '.' + data['killsub'] + chr(4)
    if data.get('double'):
        cmt = ' '.join(cmt) + ' настоящий документ признан утратившим силу повторно'
    else:
        cmt = ' '.join(cmt) + ' настоящий документ признан утратившим силу'
    # killdate = sy.replace_int_date_to_str_date(data['killdate'].strftime('%d.%m.%Y'))
    killdate = data['killdate']
    if killdate:
        cmt += ' с ' + replace_int_date_to_str_date(killdate) + ' г.'
    # print(cmt)
    return cmt


def generate_fus_cmt(data):
    kill_action = {
        'Постановление': 'признано',
        'Распоряжение': 'признано',
        'Решение': 'признано',
        'Приказ': 'признан',
        'Указ': 'признан',
        'Закон': 'признан'
    }
    cmt0 = "!STYLE J 1 72 1\nНастоящий документ фактически прекратил действие\n!STYLE J 1 72 1\n"
    cmt1 = data['killname'].split()
    cmt1[0] = chr(4) + cmt1[0] + chr(4) + data['killtopic'] + '.' + data['killsub'] + chr(4)
    cmt1 = ' '.join(cmt1)

    cmt2 = data['killed_name'].split()
    cmt2.append(kill_action.get(cmt2[0], 'признан'))
    cmt2[0] = chr(4) + cmt2[0] + chr(4) + data['killed_topic'] + chr(4)
    cmt2.append('утратившим силу')
    cmt2 = ' '.join(cmt2)

    killdate = data['killdate']
    if killdate:
        cmt2 += ' с ' + replace_int_date_to_str_date(killdate) + ' г.'
    return cmt0 + cmt1 + ' ' + cmt2


def clearending(text):
    textlist = text.split(' ')

    def delend(pattern, txt):
        try:
            rsm = re.search(pattern, txt)
            if rsm:
                txt = re.sub(pattern, '', txt)
                return txt
            else:
                return None
        except Exception:
            print('ERROR: ' + txt + ':' + pattern)

    info = ''
    i = -1
    while i < len(textlist) - 1:
        i = i + 1
        tmp = delend('ать$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ять$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('оть$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('еть$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('уть$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ешь$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ишь$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ете$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ите$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ала$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('яла$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('али$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('яли$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ола$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ела$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('оли$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ели$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ула$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ули$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ами$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('еми$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('емя$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ими$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ого$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            info = 'прилагательное'
            continue
        tmp = delend('ому$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            info = 'прилагательное'
            continue
        tmp = delend('умя$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue

        tmp = delend('ых$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue

        tmp = delend('ах$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue

        tmp = delend('ее$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ой$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ие$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('йй$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ия$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ое$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ые$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ый$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            info = 'прилагательное'
            continue
        tmp = delend('ем$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('им$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ет$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ит$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ут$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ют$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ят$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ал$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ял$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ол$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ел$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ул$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ам$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ас$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('am$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ax$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ая$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ее$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ей$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ex$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ею$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ех$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ие$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ий$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('их$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ию$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ми$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('мя$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ов$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ой$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ом$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ою$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('cm$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ум$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ух$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ую$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('шь$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('а$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('я$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('о$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('е$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ь$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ы$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('у$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('ю$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('и$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('м$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('и$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
        tmp = delend('а$', textlist[i])
        if tmp != None:
            textlist[i] = tmp
            continue
    text = ' '.join(textlist)
    return text


def clearcode(text):
    S = {'E': 'Е', 'T': 'Т', 'I': '1', 'O': '0', 'P': 'Р', 'A': 'А', 'H': 'Н', 'K': 'К', 'X': 'Х', 'C': 'С', 'B': 'В',
         'M': 'М', 'О': '0', 'З': '3'}
    text = text.upper()
    text = re.sub('[;|:|\.|,|\s|"]+$', '', text, re.IGNORECASE)
    if text.find('(') == -1:
        text = text.replace(')', '')
    ts = list(text)
    i = -1
    while i < len(ts) - 1:
        i = i + 1
        T = S.get(ts[i])
        if T != None:
            ts[i] = T
    text = ''.join(ts)
    return text


def cleartext(text):
    # text=text.upper()
    text = text.replace(chr(255), ' ')
    text = text.replace(chr(160), ' ')
    text = text.strip()
    text = re.sub('[\x01|\x02|\x03|\x05|\x06|\x07]+', '', text, re.IGNORECASE)
    tt = text.split()
    text = ''
    for t in tt:
        if t.strip() == '':
            continue
        text = text + t + ' '
    text = re.sub('\([^\)]+\)$', '', text, re.IGNORECASE)
    return text


def replaceallstrdate(lines):
    pattern = '(\d+)[\xff\s]+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)[\xff\s]+(\d\d\d\d)'
    for rsm in re.finditer(pattern, lines, re.IGNORECASE):
        if rsm.group(2).lower() == 'января':
            month = '.01.'
        if rsm.group(2).lower() == 'февраля':
            month = '.02.'
        if rsm.group(2).lower() == 'марта':
            month = '.03.'
        if rsm.group(2).lower() == 'апреля':
            month = '.04.'
        if rsm.group(2).lower() == 'мая':
            month = '.05.'
        if rsm.group(2).lower() == 'июня':
            month = '.06.'
        if rsm.group(2).lower() == 'июля':
            month = '.07.'
        if rsm.group(2).lower() == 'августа':
            month = '.08.'
        if rsm.group(2).lower() == 'сентября':
            month = '.09.'
        if rsm.group(2).lower() == 'октября':
            month = '.10.'
        if rsm.group(2).lower() == 'ноября':
            month = '.11.'
        if rsm.group(2).lower() == 'декабря':
            month = '.12.'
        if len(rsm.group(1)) == 1:
            day = '0' + rsm.group(1)
        else:
            day = rsm.group(1)
        year = rsm.group(3)
        lines = re.sub(pattern, day + month + year, lines, 1, re.IGNORECASE)
    return lines


def replace_int_date_to_str_date(text):
    month = {'01': 'января', '02': 'февраля', '03': 'марта', '04': 'апреля', '05': 'мая',
             '06': 'июня', '07': 'июля', '08': 'августа', '09': 'сентября', '10': 'октября', '11': 'ноября',
             '12': 'декабря'}
    rsm = re.search('(\d\d)[\.|/](\d\d)[\.|/](\d\d\d\d)', text)
    text = re.sub('(\d\d)[\.|/](\d\d)[\.|/](\d\d\d\d)', rsm[1].lstrip('0') + ' ' + month[rsm[2]] + ' ' + rsm[3], text)
    return text


def getmainbytopic(topic):
    inttopic = int(topic)
    result = None
    for key in eta_main.keys():
        region = eta_main[key]
        dia = region['dia']
        for d in dia:
            if inttopic >= d[0] and inttopic <= d[1]:
                result = key
                break
    return result


def getmainbysource(source):
    source = source.split('\\')
    source = source[0]
    if source == 'ОРГАНЫ СУДЕБНОЙ ВЛАСТИ РФ И СССР' or source == 'КОМПАНИЯ "ГАРАНТ"' or source == 'ГОСУДАРСТВЕННЫЕ ФОНДЫ' \
            or source == 'ОРГАНЫ ПРОКУРАТУРЫ' or source == 'СРЕДСТВА МАССОВОЙ ИНФОРМАЦИИ (СМИ)':
        return 'skip'
    result = None
    for key in eta_main.keys():
        region = eta_main[key]
        srcparent = region['srcparent']
        if srcparent == source:
            result = key
            break
    return result


def getmainbypubl(publ):
    publ = publ.split('\\')
    publ = publ[1]
    # if source=='ОРГАНЫ СУДЕБНОЙ ВЛАСТИ РФ И СССР':
    #     return 'skip'
    result = None
    for key in eta_main.keys():
        region = eta_main[key]
        publparent = region['publparent']
        if publparent == publ:
            result = key
            break
    return result


def getmainbybelongs(bel):
    if bel.find('\\') == -1:
        bel = [bel]
    else:
        bel = bel.split('\\')
    # if source=='ОРГАНЫ СУДЕБНОЙ ВЛАСТИ РФ И СССР':
    #     return 'skip'
    result = []
    for b in bel:
        if b == 'RESTRICT':
            continue
        for key in eta_main.keys():
            region = eta_main[key]
            belongs = region['belongs']
            tmpresult = None
            for b1 in belongs:
                if b1 == b:
                    tmpresult = key
                    break
            if tmpresult != None:
                break
        if tmpresult == None:
            return f'ОШИБКА\tНЕИЗВЕСТНЫЙ BELONGS {b}'
        else:
            if not tmpresult in result:
                result.append(tmpresult)
    if len(result) > 1:
        return f'ОШИБКА\tДОКУМЕНТ ИМЕЕТ BELONGS РАЗНЫХ ИБ: {", ".join(result)}'
    else:
        return ''.join(result)
    return result


def crop_garant_name(text):
    text = text[text.find('"О'):]
    text = re.sub(
        '\((утратил[а-я]*\s+силу|не\s+вступил[а-я]*\s+в\s+силу|с\s+изменениями\s+и\s+дополнениями|отменен|не\s+действует|документ\s+утратил\s+силу|прекратило\s+действие)\)',
        '', text)
    return text


def checkfatalerror(file):
    def checkheader(dinfo):
        publmain = None
        reltopicmain = None
        verltopicmain = None
        publmain = None
        belmain = None
        # получим топик
        try:
            dtype = None
            topic = '0'
            topicmain = None
            if dinfo.get('TOPIC'):
                topic = re.search('\d+', dinfo.get('TOPIC')[0])
                topic = topic[0]
                dinfo['TOPIC'] = topic
                dtype = 'DOC'
            elif dinfo.get('OBJTOPIC'):
                topic = re.search('\d+', dinfo.get('OBJTOPIC')[0])
                topic = topic[0]
                dtype = 'OBJ'
                dinfo['OBJTOPIC'] = topic
            elif dinfo.get('ANNOTOPIC'):
                topic = re.search('\d+', dinfo.get('ANNOTOPIC')[0])
                topic = topic[0]
                dtype = 'ANNO'
                dinfo['ANNOTOPIC'] = topic
            else:
                print('НЕ НАШЛИ ТОПИК', dinfo)
                exit(1)
            topicmain = getmainbytopic(topic)
        except:
            print(dinfo)
            exit(1)
        contents[topic] = dinfo['FILE']
        rel = dinfo.get('RELATED')
        if rel:
            related[topic] = rel[0]
            if len(rel) > 1:
                msg = f'{topic}\tОШИБКА\tУ ДОКУМЕНТА НЕ М.Б. БОЛЕЕ 1 СПРАВКИ (!RELATED): {", ".join(rel)}'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                deletelist.extend(rel)
                print(msg)
            else:
                reltopicmain = getmainbytopic(rel[0])
        verl = dinfo.get('VERLINK')
        if verl and verl[0] != '':
            verlink[topic] = verl[0]
            if len(verl) > 1:
                msg = f'{topic}\tОШИБКА\tУ ДОКУМЕНТА НЕ М.Б. БОЛЕЕ 1 ПРЕДЫДУЩЕЙ РЕДАКЦИИ (!VERLINK): {", ".join(verl)}'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                deletelist.extend(verl)
                print(msg)
            else:
                verltopicmain = getmainbytopic(verl[0])

        if 'EDITION' in dinfo.keys():
            dtype = 'EDITION'
        elif 'NODOC' in dinfo.keys() or 'REFDOC' in dinfo.keys():
            dtype = 'NODOC'

        if dtype != 'ANNO':
            belmain = None
            value = dinfo.get('BELONGS')
            if not value:
                msg = f'{topic}\tОШИБКА\tНЕТ BELONGS'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            else:
                value = '\\'.join(value)
                belmain = getmainbybelongs(value.upper())
                # print(belmain)
                if belmain.find('ОШИБКА') > -1:
                    msg = topic + '\t' + belmain
                    errors.extend(['ROW', msg])
                    deletelist.append(topic)
                    print(msg)

        srcmain = None
        if dtype == 'DOC':
            if dinfo.get('SOURCE') == None:
                msg = f'{topic}\tОШИБКА\tНЕТ ОРГАНА'
                deletelist.append(topic)
                errors.extend(['ROW', msg])
                print(msg)
            else:
                for src in dinfo.get('SOURCE'):
                    tempsrcmain = getmainbysource(src.upper())
                    if tempsrcmain == None:
                        msg = f'{topic}\tОШИБКА\tУКАЗАН НЕ РЕГИОНАЛЬНЫЙ ОРГАН: {src.upper()}'
                        errors.extend(['ROW', msg])
                        deletelist.append(topic)
                        print(msg)
                    elif tempsrcmain == 'skip':
                        continue
                    else:
                        if srcmain != None and srcmain != tempsrcmain:
                            msg = f'{topic}\tОШИБКА\tУКАЗАНЫ ОРГАНЫ РАЗНЫХ ИБ: {srcmain} <> {tempsrcmain}'
                            errors.extend(['ROW', msg])
                            deletelist.append(topic)
                            print(msg)
                        srcmain = tempsrcmain

            if dinfo.get('PUBLISHEDIN'):
                for publ in dinfo.get('PUBLISHEDIN'):
                    temppublmain = getmainbypubl(publ.upper())
                    if temppublmain == None:
                        msg = f'{topic}\tОШИБКА\tУКАЗАНА НЕ РЕГИОНАЛЬНАЯ ПУБЛИКАЦИЯ: {publ.upper()}'
                        errors.extend(['ROW', msg])
                        deletelist.append(topic)
                        print(msg)
                    elif temppublmain == 'skip':
                        continue
                    else:
                        if publmain != None and publmain != temppublmain:
                            msg = f'{topic}\tОШИБКА\tУКАЗАНЫ ПУБЛИКАЦИИ РАЗНЫХ ИБ: {publmain} <> {temppublmain}'
                            errors.extend(['ROW', msg])
                            deletelist.append(topic)
                            print(msg)
                        publmain = temppublmain

            if dinfo.get('TYPE') == None:
                msg = f'{topic}\tОШИБКА\tНЕТ ТИПА'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)

        value = dinfo.get('MAIN')
        if not value:
            msg = f'{topic}\tОШИБКА\tНЕТ MAIN'
            errors.extend(['ROW', msg])
            deletelist.append(topic)
            print(msg)
        else:
            if len(value) > 1:
                msg = f'{topic}\tОШИБКА\tУ ДОКУМЕНТА НЕ М.Б. БОЛЕЕ 1 MAIN: {", ".join(value)}'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            elif len(value) == 1:
                value = ''.join(value)
                region = eta_main.get(value)
                if region == None:
                    msg = f'{topic}\tОШИБКА\tНЕИЗВЕСТНЫЙ MAIN "{value}", НЕЛЬЗЯ ОПРЕДЕЛИТЬ БЛОК ИБ'
                    errors.extend(['ROW', msg])
                    deletelist.append(topic)
                    print(msg)
            if topicmain and topicmain != value:
                msg = f'{topic}\tОШИБКА\tТОПИК: "{topicmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            if reltopicmain and reltopicmain != value:
                msg = f'{topic}\tОШИБКА\tRELATED: "{reltopicmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            if verltopicmain and verltopicmain != value:
                msg = f'{topic}\tОШИБКА\tVERLINK: "{verltopicmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            if publmain and publmain != value:
                msg = f'{topic}\tОШИБКА\tПУБЛИКАЦИЯ: "{publmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            if srcmain and srcmain != value:
                msg = f'{topic}\tОШИБКА\tОРГАН: "{srcmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
            if belmain and belmain != value:
                msg = f'{topic}\tОШИБКА\tГРУППА ДОКУМЕНТА BELONGS: "{belmain}" НЕ СООТВЕТСТВУЕТ MAIN: "{value}"'
                errors.extend(['ROW', msg])
                deletelist.append(topic)
                print(msg)
        docinfo.append(dinfo)
        # return result

    result = []
    docinfo = []
    verlink = {}
    related = {}
    contents = {}
    errors = []
    deletelist = []
    if file.lower().find('106-112.nsr') > -1:
        return []
    with open(file, 'r', encoding='cp866') as fl:
        lines = fl.readlines()
    docsize = 0
    dinfo = {}
    # dinfo['FILE']=file
    stylecount = 0
    topiccount = 0
    skip = False
    buff = []
    for line in lines:
        if line.strip() == '':
            continue
        elif line[0] == ';':
            continue
        elif line[0] == '!':
            if len(buff) > 0:
                buff = ' '.join(buff)
                rsm = re.findall('\x04([^\x04]+)\x04([^\x04]+)\x04', buff)
                for r in rsm:
                    # print(r)
                    lnk = r[1]

                    if lnk.find('.') > -1:
                        lnk = lnk[:lnk.find('.')]
                    if lnk == '111':
                        msg = f'{ttop[0]}\tОШИБКА\tССЫЛКА НА 111: {buff}'
                        errors.extend(['ROW', msg])
                        deletelist.append(ttop[0])
                buff = []

            rsm = re.search('^!([A-Z]+)(.*)', line)
            if rsm:
                if rsm[1] in ['TOPIC', 'ANNOTOPIC', 'OBJTOPIC']:
                    if len(dinfo) > 0:
                        dinfo['FILE'] = file
                        checkheader(dinfo)
                        dinfo = {}

                    stylecount = 0
                    docsize = 0
                    skip = False
                    topiccount = topiccount + 1
                    ttop = re.search('\d+', line)
                if rsm[1] in ['STYLE', 'SUB', 'BLOCK']:
                    skip = True
                if line.strip().upper() == '!TYPE ФЕДЕРАЛЬНЫЙ ЗАКОН':
                    msg = f'{ttop[0]}\tОШИБКА\tУ НАС НЕ БЫВАЕТ ФЕДЕРАЛЬНЫХ ЗАКОНОВ: {line}'
                    errors.extend(['ROW', msg])
                    deletelist.append(ttop[0])
                    print()
                if rsm[1] == '!STYLE':
                    stylecount += 1
                elif skip == False:
                    value = dinfo.get(rsm[1])
                    if value == None:
                        value = [rsm[2].strip()]
                    else:
                        value.append(rsm[2].strip())
                    dinfo[rsm[1]] = value
        else:
            docsize += len(line)
            buff.append(line.strip())
    if len(dinfo) > 0:
        dinfo['FILE'] = file
        checkheader(dinfo)
    print(f'{len(result)}, topics: {topiccount} in {file}')
    return [docinfo, contents, verlink, related, errors, deletelist]


def saveasjson(dict, path):
    with open(path, 'w', encoding='utf-8') as fl:
        json.dump(dict, fl, ensure_ascii=False, indent=4)
        print('save:', path)


def getstatnsr(file, **kwargs):
    result = []
    topic = ''
    dmain = ''
    dtype = ''
    dnorm = ['О']
    verlink = ''
    related = ''
    vano = []
    vinc = []
    vold = []
    vch = []
    buff = []
    stylecount = 0
    index = 0
    size = 0
    with open(file, 'r', encoding='cp866') as fl:
        lines = fl.readlines()
    for line in lines:
        if line.strip() == '':
            continue
        elif line[0] == ';':
            continue
        elif line[0] == '!':
            if len(buff) > 0:
                buff = ' '.join(buff)
                buff = re.sub('\x04([^\x04]+)\x04([^\x04]+)\x04', '\g<1>', buff)
                buff = re.sub('\s{2,}', ' ', buff)
                buff = re.sub('[\x01|\x02|\x03|\x04|\x05|\x06|\x07]', '', buff)
                size += len(buff) + 1
                # rsm=re.findall('\S+',line)
                # if rsm:
                #     for r in rsm:
                #         size+=len(r)+1
                buff = []
            if line.strip() == '!REL':
                dtype = 'REL'
            elif line.strip() == '!EDITION':
                dtype = 'EDITION'
            elif line.strip().upper() == '!NORM ИНДИВИДУАЛЬНЫЕ':
                dnorm = ['И']
            rsm = re.search('^(!\S+)\s+(\S+)', line)
            if rsm:
                if index > 0 and rsm[1] in ['!TOPIC', '!ANNOTOPIC', '!OBJTOPIC']:
                    result.append(
                        [topic, dmain, dtype, dnorm, verlink, related, vano, vinc, vold, vch, stylecount, [size]])
                    topic = ''
                    dmain = ''
                    dtype = ''
                    dnorm = ['О']
                    verlink = ''
                    related = ''
                    vano = []
                    vinc = []
                    vold = []
                    vch = []
                    stylecount = 0
                    size = 0
                if rsm[1] == '!TOPIC':
                    dtype = 'DOC'
                    topic = rsm[2]
                    index += 1
                elif rsm[1] == '!ANNOTOPIC':
                    dtype = 'ANNO'
                    topic = rsm[2]
                    index += 1
                elif rsm[1] == '!OBJTOPIC':
                    dtype = 'OBJ'
                    topic = rsm[2]
                    index += 1
                elif rsm[1] == '!RELATED':
                    related = rsm[2]
                elif rsm[1] == '!VERLINK':
                    verlink = rsm[2]
                elif rsm[1] == '!MAIN':
                    dmain = rsm[2]
                elif rsm[1] == '!VANONCED':
                    vano.append('-'.join(reversed(rsm[2].split('/'))))
                elif rsm[1] == '!VINCLUDED':
                    vinc.append('-'.join(reversed(rsm[2].split('/'))))
                elif rsm[1] == '!VCHANGED':
                    vch.append('-'.join(reversed(rsm[2].split('/'))))
                elif rsm[1] == '!VABOLISHED':
                    vold.append('-'.join(reversed(rsm[2].split('/'))))
                elif rsm[1] == '!STYLE':
                    if rsm:
                        stylecount += 1
        else:
            buff.append(line.strip())

    if index > 0:
        result.append([topic, dmain, dtype, dnorm, verlink, related, vano, vinc, vold, vch, stylecount, [size]])
    print(f'{len(result)} in {file}')
    if len(result) > 0:
        result1 = [file, result]
        return result1
    else:
        return [file, []]


def extracttopicfromfile(topic, file, targetfolder):
    def save(buff, path):
        try:
            print(f'save file: {path}')
            with open(path, 'w', encoding='cp866') as fl:
                fl.writelines(buff)
            print('ok')
        except Exception as e:
            print('ERROR', e)

    with open(file, 'r', encoding='cp866') as fl:
        lines = fl.readlines()
    buff = []
    writing = False
    prefix = None
    main = 'UNC'
    if targetfolder[-1] != '\\':
        targetfolder += '\\'
    for line in lines:
        rsm = re.search('^!MAIN (\S+)', line)
        if rsm:
            main = rsm[1]
        rsm = re.search('^(![TOPIC|ANNOTOPIC|OBJTOPIC]+)\s+(\d+)', line)
        if rsm:
            if len(buff) > 0:
                save(buff, f'{targetfolder}{main}-{topic}-{prefix}.nsr')
                writing = False
                buff = []
            if rsm[1] == '!TOPIC':
                prefix = 'd'
            elif rsm[1] == '!ANNOTOPIC':
                prefix = 'a'
            elif rsm[1] == '!OBJTOPIC':
                prefix = 'o'
            main = 'UNC'
            if rsm[2] == topic:
                writing = True
        if writing == True:
            buff.append(line)
    if len(buff) > 0:
        save(buff, f'{targetfolder}{main}-{topic}-{prefix}.nsr')


def extenddict(target, source):
    if target == None:
        target = {}
    for key in source.keys():
        if not key in target.keys():
            target[key] = [source[key]]
        else:
            target[key].append(source[key])
    return target


def summdict(dict1, dict2):
    if dict1 == None:
        dict1 = {}
    for key in dict2.keys():
        if not key in dict1.keys():
            dict1[key] = dict2[key]
        else:
            dict1[key] += dict2[key]
    return dict1


class NsrDoc():
    def __init__(self, lines):
        self.doc = lines
        self.subs = set()

    def loadfromfile(self, path):
        with open(path, 'r', encoding='cp866') as fl:
            self.doc = fl.readlines()

    def saveasfile(self, path):
        with open(path, 'w', encoding='cp866') as fl:
            fl.writelines(self.doc)

    def clear(self):
        del (self.doc)
        self.doc = []

    def getcmd(self, cmd):
        result = []
        for i in range(0, len(self.doc)):
            rsm = re.search('^' + cmd + '\s*(.*)', self.doc[i])
            if rsm:
                result.append([i, cmd, rsm[1].strip()])
        if result == []:
            result = None
        return result

    def gettopic(self):
        # print('gettopic')
        result = None
        for i in range(0, len(self.doc)):
            rsm = re.search('^!(ANNOTOPIC|OBJTOPIC|TOPIC)\s+(\d+)', self.doc[i])
            if rsm:
                # print(rsm[)
                return [rsm[2], rsm[1]]
        return result

    def delcmd(self, cmd):
        for i in reversed(range(0, len(self.doc))):
            rsm = re.search('^' + cmd + '\s*(.*)', self.doc[i])
            if rsm:
                self.doc.pop(i)

    def insertcmdafter(self, cmd, value):
        for i in range(0, len(self.doc)):
            rsm = re.search('^' + cmd + '\s*(.*)', self.doc[i])
            if rsm:
                self.doc.insert(i + 1, value + '\n')
                break

    def get_size(self):
        result = 0
        for l in self.doc:
            result += len(l)
        return result

    def get_subs(self):
        for line in self.doc:
            rsm = re.search('!(BLOCK|SUB)\s+(\d+)', line)
            if rsm:
                self.subs.add(rsm[2])
        print(self.subs)

    # прочесть таблицу в список с позиции
    # вернуть позицию начала, конца и лист со строками таблицы
    def read_table(self, from_pos):
        flag = 0
        tablestart = 0
        tablestop = 0
        result = []
        for i in range(from_pos, len(self.doc)):
            if flag == 2:
                break
            if self.doc[i].find('!TABLE ') > -1:
                flag = 1
                result.append(self.doc[i])
                tablestart = i
            elif self.doc[i].find('!TABLEEND') > -1:
                flag = 2
                tablestop = i
                result.append(self.doc[i])
            elif flag == 1:
                result.append(self.doc[i])
        if flag == 2:
            return (tablestart, tablestop, result)

    def read_block(self, from_pos):
        level = 0
        result = []
        for i in range(from_pos, len(self.doc)):
            if re.search('^!BLOCK\s+',self.doc[i]):
                level += 1
            elif re.search('^!BLOCKEND',self.doc[i]):
                level -= 1
            result.append(self.doc[i])
            if level == 0:
                break
        return result
#


class NsrFile():
    def __init__(self, path):
        self.path = path
        self.lines = []
        self.pathfilepath = os.path.split(self.path)[0] + '\\pathfile.txt'

    def close(self):
        del (self.lines)

    def read(self, **kwargs):
        with open(self.path, 'r', encoding='cp866', errors='ignore') as fl:
            lines = fl.readlines()
        pathfile = open(self.pathfilepath, 'a', encoding='cp866', errors='ignore')
        self.docs = []
        doc = []
        tcount = 0
        index = -1
        topic = ''
        writed_topic = False
        while True:
            index += 1
            if index == len(lines):
                self.docs.append(doc)
                doc = []
                break
            line = lines[index]
            if kwargs.get('ignore_tech_comments'):
                if len(line) > 0 and line[0] == ';':
                    if not writed_topic:
                        pathfile.write(topic)
                    pathfile.write(str(index) + '\t' + line)
                    continue
                elif line[0] == '!' and line.find('!*NAMECOMMENT') > -1:
                    if not writed_topic:
                        pathfile.write(topic)
                    pathfile.write(str(index) + '\t' + line)
                    continue
            if kwargs.get('ignore_alarms') and line.find('!ALARMS') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if kwargs.get('ignore_main') and line.find('!MAIN') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if kwargs.get('ignore_log') and line.find('!*LOG') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if kwargs.get('ignore_stage') and line.find('!STAGE') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if kwargs.get('ignore_division') and line.find('!DIVISION') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if kwargs.get('ignore_izm') and line.find('!IZM') > -1:
                if not writed_topic:
                    pathfile.write(topic)
                pathfile.write(str(index) + '\t' + line)
                continue
            if line[:6] == '!TOPIC' or line[:10] == '!ANNOTOPIC' or line[:9] == '!OBJTOPIC':
                tcount += 1
                writed_topic = False
                topic = line
                if tcount > 1:
                    self.docs.append(doc)
                    doc = []
            doc.append(line)

        self.docscount = len(self.docs)
        # print(tcount)
        del (lines)
        pathfile.close()

    def save(self):
        with open(self.path, 'w', encoding='cp866', errors='ignore') as fl:
            for doc in self.docs:
                fl.writelines(doc)

    def saveas(self, newpath):
        with open(newpath, 'w', encoding='cp866', errors='ignore') as fl:
            for doc in self.docs:
                fl.writelines(doc)

    def saveasjson(self):
        saveasjson(self.docs, self.path + '.json')

    def split(self, size):
        size = size * 1024 * 1024
        tempsize = 0
        index = 1
        partfile, fileext = os.path.splitext(self.path)
        tempfile = open(partfile + f'-{str(index)}' + fileext, 'w', encoding='cp866', errors='ignore')
        for doc in self.docs:
            for line in doc:
                tempsize += len(line)
            tempfile.writelines(doc)
            tempfile.write('\n')
            if tempsize > size:
                tempsize = 0
                tempfile.close()
                index += 1
                tempfile = open(partfile + f'-{str(index)}' + fileext, 'w', encoding='cp866', errors='ignore')
        tempfile.close()

    def splitbydoc(self):
        folder = os.path.split(self.path)[0] + '\\'
        for doc in self.docs:
            prefix = ''
            tempdoc = NsrDoc(doc)
            id, topic = tempdoc.gettopic()
            # print(id)
            if topic == 'OBJTOPIC':
                prefix = 'o'
            elif topic == 'ANNOTOPIC':
                prefix = 'a'
            with open(folder + id + prefix + '.nsr', 'w', encoding='cp866', errors='ignore') as fl:
                fl.writelines(doc)
            del (tempdoc)

    def deletefile(self):
        os.rename(self.path, self.path.replace('.nsr', '.bak'))


class FilterStatNsrc():
    def __init__(self, statfolder, **kwargs):
        self.path = statfolder
        self.statistic = {}
        self.contents = {}
        self.supermap = {}
        if self.path[-1] != '\\':
            self.path += '\\'
        self.fromdate = kwargs.get('fromdate')
        self.date = kwargs.get('todate')

    def loadstatistic(self):
        if os.path.exists(self.path + 'statistic.json'):
            with open(self.path + 'statistic.json', 'r', encoding='utf-8') as fh:  # открываем файл на запись
                self.statistic = json.load(fh)
            print(len(self.statistic))

    def loadcontents(self):
        if os.path.exists(self.path + 'contents.json'):
            with open(self.path + 'contents.json', 'r', encoding='utf-8') as fh:  # открываем файл на запись
                self.contents = json.load(fh)
            print(len(self.contents))

    def loadsupermap(self):
        if os.path.exists(self.path + 'supermap.json'):
            with open(self.path + 'supermap.json', 'r', encoding='utf-8') as fh:  # открываем файл на запись
                self.supermap = json.load(fh)

    def getfulltopiclist(self):
        if len(self.contents) == 0:
            self.loadcontents()
        return self.contents.keys()

    def getdocsinfo(self, topics, cols):
        result = []
        for topic in topics:
            rsm = re.search('\d+', topic)
            row = []
            if rsm:
                print(rsm[0])
                row.append(rsm[0])
                info = self.contents.get(rsm[0], None)
                print(info)
                if info:
                    if 'DTYPE' in cols:
                        row.append(info.get('dtype', '[НЕ ОПРЕДЕЛЕНО]'))
                    if 'DMAIN' in cols:
                        row.append(info.get('dmain', '[НЕ ОПРЕДЕЛЕНО]'))
                    if 'VERLINK' in cols:
                        row.append(info.get('verlink', '0'))
                    if 'FILE' in cols:
                        row.append(info.get('file', '[НЕ ОПРЕДЕЛЕНО]'))
                else:
                    row.append('[НЕТ ДАННЫХ]')
            result.append('\t'.join(row))
        return result

    def getTopRelVerlFromContents(self, topics):
        if len(self.contents) == 0:
            self.loadcontents()
        topics = topics.split(' ')
        result = {}
        for topic in topics:
            if topic.strip() == '':
                continue
            r = self.contents.get(topic)
            if r:
                result[topic] = r
                print(topic, 'ok')
                temp = r.get('verlink')
                if temp:
                    topics.append(temp)
                temp = r.get('related')
                if temp:
                    topics.append(temp)
            else:
                print(f'отсутствует топик {topic}')
        if len(result) > 0:
            return result

    def gettopicsfromsupermap(self, **kwargs):
        report = []
        for doc in self.supermap:
            topic, dmain, dtype, dnorm, verlink, related, vano, vinc, vold, vch, stylecount, file = doc
            param = kwargs.get('main')
            if param:
                if not param == dmain:
                    continue
            param = kwargs.get('vinc')
            if param:
                if len(vinc) == 0:
                    continue
                if vinc[0].find(param) == -1:
                    continue
            report.append(f'{topic}\t{dtype}\t{vinc[0]}\n')
        with open(self.path + '\\gettopicsfromsupermap.txt', 'w', encoding='utf-8') as fl:
            fl.writelines(sorted(report))
        print('ok')

    def generatestatreport(self):
        if len(self.statistic) == 0:
            self.loadstatistic()
            print(len(self.statistic))
        self.totaldict = {}
        bymonthdict = {}
        for key in reversed(sorted(self.statistic.keys())):
            if key == 'relcount':
                continue
            subkey = '-'.join(key.split('-')[:-1])
            # запрос из кучи данных. Здесь данные за день. По регионам
            regionz = self.statistic[key]
            # перебор регионов из кучи данных
            # регкей это белонгс
            for regkey in regionz.keys():
                # переберем MAIN из порции за дату
                info = regionz[regkey]
                # в инфо находится статистика за день в одном блоке
                reg_rep_info = bymonthdict.get(regkey)
                if not reg_rep_info:
                    reg_rep_info = {}
                cur_stat_info = reg_rep_info.get(subkey)
                if not cur_stat_info:
                    cur_stat_info = info
                else:
                    for k in info.keys():
                        cur_stat_info[k] += info[k]
                reg_rep_info[subkey] = cur_stat_info
                bymonthdict[regkey] = reg_rep_info
        return bymonthdict

    def getstatfortelegramm(self, plain):
        self.loadstatistic()
        stinfo = self.generatestatreport()
        result = ''
        for key in stinfo.keys():
            if key.strip() == '':
                continue
            yeardict = {}
            monthdict = {}
            percentresult = []
            yearname = None
            result += f'\nРегион: {key}\n'
            info = stinfo[key]
            ydoccount = 0
            mdoccount = 0
            for i in info:
                if yearname == None:
                    yearname = i.split('-')[0]
                    ydoccount = 0
                    mdoccount = 0
                    if plain:
                        ydoccount = plain.get(key + '-' + yearname)
                        if ydoccount:
                            mdoccount = ydoccount / 12
                        else:
                            mdoccount = 0
                if len(monthdict) == 0:
                    monthdict = info
                    result += f'за текущий месяц: {i}\n'
                    result += f"подключено док: {info[i]['vinc']} шт."
                    if mdoccount > 0:
                        result += f' ({mdoccount})'
                        result += f"\nплан за месяц выполнен на {round(info[i]['vinc'] * 100 / mdoccount, 2)}%"
                    result += f"\nподключено ред. {info[i]['edition']} шт., изменено: {info[i]['vch']} шт., утрат {info[i]['vold']} шт.\n"
                if len(yeardict) == 0:
                    yeardict = info[i]
                else:
                    yeardict = summdict(yeardict, info[i])
                if i.find('-12') != -1:
                    result += f'\nза текущий год: {yearname}\n'
                    result += f"подключено док: {yeardict['vinc']} шт."
                    if ydoccount != None and ydoccount > 0:
                        result += f' (план: {ydoccount})'
                        result += f"\nплан выполнен на {round(yeardict['vinc'] * 100 / ydoccount, 2)}%"
                    result += f"\nподключено ред.{yeardict['edition']} шт., изменено: {yeardict['vch']} шт., утрат {yeardict['vold']} шт.\n"
                    result += '-' * 10 + '\n'
                    yeardict = {}
                    break
            if len(yeardict) > 0:
                result += f'\nза текущий год: {yearname}\n'
                result += f"подключено док: {yeardict['vinc']} шт."
                if ydoccount != None and ydoccount > 0:
                    result += f' (план: {ydoccount})'
                    result += f"\nплан выполнен на {round(yeardict['vinc'] * 100 / ydoccount, 2)}%"
                result += f"\nподключено ред.{yeardict['edition']} шт., изменено: {yeardict['vch']} шт., утрат {yeardict['vold']} шт.\n"
                yeardict = {}
        return result

    def saveasXLS(self, dict, path):
        sheets = {}
        colnames = ['период', 'кол-во всех док-в', 'средний вес', 'вес всех док-в', 'из них инд.', 'вес инд. док-в',
                    'редакции', 'изменения', 'утрата']
        wb = opx.Workbook()
        wb.remove(wb.active)
        wb.add_named_style(tablehead)
        print('keys:', dict.keys())
        for key in dict.keys():
            if key.strip() == '':
                continue
            sheets[key] = wb.create_sheet(key)
            yeardict = {}
            r = 2
            c = 1
            print(key)
            for cn in colnames:
                sheets[key].cell(row=r, column=c, value=cn)
                if r == 2 and c == 1:
                    sheets[key].cell(row=r, column=c, value=cn).style = tablecolhead
                else:
                    sheets[key].cell(row=r, column=c, value=cn).style = tablehead
                column_letter = get_column_letter((c))
                sheets[key].column_dimensions[column_letter].width = 25
                c += 1
            info = dict[key]
            yearname = None
            c = 1
            for i in info:
                print(i)
                r = r + 1
                if yearname == None:
                    yearrow = r
                    r += 1
                    yearname = i.split('-')[0]
                    sheets[key].cell(row=yearrow, column=1, value=f'Всего за {yearname}').style = tablecolhead
                sheets[key].cell(row=r, column=1, value=i)
                sheets[key].cell(row=r, column=2, value=info[i]['vinc'])
                if info[i]['vinc'] != 0:
                    sheets[key].cell(row=r, column=3, value=round(info[i]['size'] / 1024 / info[i]['vinc'], 2))
                sheets[key].cell(row=r, column=4, value=round(info[i]['size'] / 1024, 2))
                sheets[key].cell(row=r, column=5, value=info[i]['dnorm'])
                sheets[key].cell(row=r, column=6, value=round(info[i]['isize'] / 1024, 2))
                sheets[key].cell(row=r, column=7, value=info[i]['edition'])
                sheets[key].cell(row=r, column=8, value=info[i]['vch'])
                sheets[key].cell(row=r, column=9, value=info[i]['vold'])
                if len(yeardict) == 0:
                    yeardict = info[i]
                else:
                    yeardict = summdict(yeardict, info[i])
                if i.find('-12') != -1:
                    sheets[key].cell(row=yearrow, column=2, value=yeardict['vinc']).style = tablecolhead
                    if yeardict['vinc'] != 0:
                        sheets[key].cell(row=yearrow, column=3, value=round(yeardict['size'] / 1024 / yeardict['vinc'],
                                                                            2)).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=4,
                                     value=round(yeardict['size'] / 1024, 2)).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=5, value=yeardict['dnorm']).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=6,
                                     value=round(yeardict['isize'] / 1024, 2)).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=7, value=yeardict['edition']).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=8, value=yeardict['vch']).style = tablecolhead
                    sheets[key].cell(row=yearrow, column=9, value=yeardict['vold']).style = tablecolhead
                    r += 2
                    yearrow = r
                    yearname = i.split('-')[0]
                    sheets[key].cell(row=yearrow, column=1, value=f'Всего за {yearname}').style = tablecolhead
                    yeardict = {}
            if len(yeardict) > 0:
                sheets[key].cell(row=yearrow, column=2, value=yeardict['vinc']).style = tablecolhead
                if yeardict['vinc'] != 0:
                    sheets[key].cell(row=yearrow, column=3,
                                     value=round(yeardict['size'] / 1024 / yeardict['vinc'], 2)).style = tablecolhead
                sheets[key].cell(row=yearrow, column=4, value=round(yeardict['size'] / 1024, 2)).style = tablecolhead
                sheets[key].cell(row=yearrow, column=5, value=yeardict['dnorm']).style = tablecolhead
                sheets[key].cell(row=yearrow, column=6, value=round(yeardict['isize'] / 1024, 2)).style = tablecolhead
                sheets[key].cell(row=yearrow, column=7, value=yeardict['edition']).style = tablecolhead
                sheets[key].cell(row=yearrow, column=8, value=yeardict['vch']).style = tablecolhead
                sheets[key].cell(row=yearrow, column=9, value=yeardict['vold']).style = tablecolhead
            print('create excel sheet', key)
        wb.save(path)
        wb.close()


class GetStatNsrc():
    def __init__(self, ):
        self.files = []
        self.report = {}
        self.processcount = cpu_count()

    def addfiles(self, path):
        if path[-1] != '\\':
            path += '\\'
        # if not os.path.exists(path):
        try:
            self.files.extend(glob.glob(path + '*.nsr'))
        except Exception as E:
            print(f'ОШИБКА: {path}')
            print(E)
            return

    # date - belongs - ver1 ver2 ver3

    def execute(self, reportfolder):
        def updateinfo(statistic, r):
            # print(statistic)
            topic, dmain, dtype, dnorm, verlink, related, vano, vinc, vold, vch, stylecount, size, file = r
            if dmain == None:
                print(topic, 'не имеет команду MAIN')
                return statistic
            versionsdata = {'vano': vano, 'vinc': vinc, 'vold': vold, 'vch': vch}
            # print(versionsdata)
            for key in versionsdata.keys():
                # print('key:',key)
                # print('value',versionsdata[key])
                for vdate in versionsdata[key]:
                    # print('vdate',vdate)
                    info = statistic.get(vdate, {})
                    subinfo = info.get(dmain,
                                       {'vano': 0, 'vinc': 0, 'vold': 0, 'vch': 0, 'edition': 0, 'dnorm': 0, 'size': 0,
                                        'isize': 0})
                    if dtype == 'DOC':
                        if key == 'vinc':
                            subinfo['size'] += size[0]
                        subinfo[key] += 1
                        if dtype == 'DOC' and dnorm == ['И']:
                            subinfo['dnorm'] += 1
                            subinfo['isize'] += size[0]
                    elif dtype == 'EDITION':
                        subinfo['edition'] += 1
                    info[dmain] = subinfo
                    statistic[vdate] = info
            return statistic

        if reportfolder[-1] != '\\':
            reportfolder += '\\'
        if not os.path.exists(reportfolder):
            try:
                os.makedirs(reportfolder)
                print(f'создана папка: {reportfolder}')
            except:
                print(f'не удалось создать папку: {reportfolder}')
        print(reportfolder)
        time.sleep(4)
        pool = Pool(processes=self.processcount)
        data = pool.map(getstatnsr, self.files)
        with open(reportfolder + 'map.json', 'w', encoding='utf-8') as fl:
            json.dump(data, fl, ensure_ascii=False, indent=4)
        sresult = []
        contents = {}
        statistic = {}
        maxdate = None
        mandate = None
        doccount = 0
        relcount = 0
        editioncount = 0
        vinccount = 0
        vchangedcount = 0
        vanoncedcount = 0
        vabolishedcount = 0
        size = 0
        deleteemptynsr = []
        anoncednotincluded = []
        for d in data:
            if d == None:
                continue
            if len(d) < 2:
                continue
            file, result = d
            if len(result) == 0:
                path, filename = os.path.split(file)
                filename = os.path.splitext(filename)
                deleteemptynsr.append(f'ren {file} {filename[0]}.zero\n')
            while len(result) != 0:
                r = result.pop(0)
                r.append(file)
                sresult.append(r)
                topic, dmain, dtype, dnorm, verlink, related, vano, vinc, vold, vch, stylecount, size, file = r
                print(r)
                if len(vinc) == 0 and len(vano) > 0:
                    anoncednotincluded.append(topic + f' анонсирован: {str(vano)}, но не подключен\n')
                r1 = {}
                r1['dtype'] = dtype
                r1['dmain'] = dmain
                # r1['size']=size
                if verlink != '':
                    r1['verlink'] = verlink
                if related != '':
                    r1['related'] = related
                if stylecount != '':
                    r1['stylecount'] = stylecount
                if file != '':
                    r1['file'] = file
                contents[topic] = r1
                statistic = updateinfo(statistic, r)
                if dtype == 'REL':
                    if not 'relcount' in statistic.keys():
                        statistic['relcount'] = 1
                    else:
                        statistic['relcount'] += 1
        print('save data...')
        if len(sresult) > 0:
            with open(reportfolder + 'supermap.json', 'w', encoding='utf-8') as fl:
                json.dump(sresult, fl, ensure_ascii=False, indent=4)
            print('save:', reportfolder + 'supermap.json')
        if len(statistic) > 0:
            with open(reportfolder + 'statistic.json', 'w', encoding='utf-8') as fl:
                json.dump(statistic, fl, ensure_ascii=False, indent=4)
            print('save:', reportfolder + 'statistic.json')
        if len(contents) > 0:
            with open(reportfolder + 'contents.json', 'w', encoding='utf-8') as fl:
                json.dump(contents, fl, ensure_ascii=False, indent=4)
            print('save:', reportfolder + 'contents.json')
        if len(deleteemptynsr) > 0:
            with open(reportfolder + 'deleteemptynsr.cmd', 'w', encoding='cp1251') as fl:
                fl.writelines(deleteemptynsr)
            print('save:', reportfolder + 'deleteemptynsr.cmd')
        if len(anoncednotincluded) > 0:
            with open(reportfolder + 'anoncednotincluded.txt', 'w', encoding='cp1251') as fl:
                fl.writelines(anoncednotincluded)
            print('save:', reportfolder + 'anoncednotincluded.txt')
        print('execute okk')


class CheckFatal():

    def __init__(self, path, report):
        self.files = []
        self.report = {}
        self.processcount = cpu_count()
        self.reportfolder, self.reportfile = os.path.split(report)
        self.reportfolder += '\\'
        self.errors = []
        self.deletelist = []

    def reportfullpath(self):
        print(self.reportfolder + self.reportfile)

    def addfiles(self, path):
        if not os.path.exists(os.path.split(path)[0]):
            print(f'ОШИБКА: указан не существующий путь: {path}')
            return
        self.files.extend(glob.glob(path))

    def execute(self):
        # Вот прям тут пошла обработка НСР многопоточная
        pool = Pool(processes=self.processcount)
        data = pool.map(checkfatalerror, self.files)
        bigdinfo = []
        bigcontents = {}
        bigverlink = {}
        bigrelated = {}
        index = 0
        for d in data:
            try:
                dinfo, contents, verlink, related, errs, dells = d
            except:
                print(d)
                continue
            bigdinfo.extend(dinfo)
            bigcontents = extenddict(bigcontents, contents)
            print('contents:', len(contents))
            bigverlink = extenddict(bigverlink, verlink)
            bigrelated = extenddict(bigrelated, related)
            self.errors.extend(errs)
            self.deletelist.extend(dells)
        r_verlink = {}
        for key in bigverlink.keys():
            value = bigverlink[key]
            if not value[0] in bigcontents.keys():
                msg = f'ОШИБКА\tдокумент ссылается на отсутствующую редакцию VERLINK {value[0]}'
                self.errors.extend(['ROW', key, msg])
                self.deletelist.append(key)
                print(['ROW', value, msg])
            if value[0] in r_verlink.keys():
                r_value = r_verlink[value[0]]
                msg = f'{key}\tдва топика ссылаются на одну редакцию: {value[0]}'
                self.errors.extend(['ROW', r_value[0], msg])
                self.deletelist.append(key)
                self.deletelist.append(r_value)
                print(['ROW', value[0], msg])
            r_verlink[value[0]] = key
        for key in bigrelated.keys():
            value = bigrelated[key]
            if not value[0] in bigcontents.keys():
                msg = f'ОШИБКА\tдокумент ссылается на отсутствующую справку RELATED {value[0]}'
                self.errors.extend(['ROW', key, msg])
                self.deletelist.append(key)
                print(['ROW', value[0], msg])

        tdict = {}
        for dinfo in bigdinfo:
            for dtype in ['TOPIC', 'OBJTOPIC', 'ANNOTOPIC']:
                if not dtype in dinfo.keys():
                    continue
                count, files = tdict.get(dtype + '\t' + dinfo[dtype], [None, None])
                if count:
                    files = files + '\t' + dinfo['FILE'].strip()
                    count += 1
                    tdict[dtype + '\t' + dinfo[dtype]] = [count, files]
                else:
                    tdict[dtype + '\t' + dinfo[dtype]] = [1, dinfo['FILE'].strip()]
        for key in tdict.keys():
            count, files = tdict.get(key)
            if count > 1:
                t1, t2 = key.split('\t')
                msg = f"ОШИБКА\tповторяющийся {t1}\t{count}\t{files}"
                self.errors.extend(['ROW', t2, msg])
                self.deletelist.append(t2)

        for dd in self.deletelist:
            t = bigrelated.get(dd)
            if t:
                if not t[0] in self.deletelist:
                    self.deletelist.append(t[0])
            t = bigverlink.get(dd)
            if t:
                if not t[0] in self.deletelist:
                    self.deletelist.append(t[0])
            t = r_verlink.get(dd)
            if t:
                if not t in self.deletelist:
                    self.deletelist.append(t)

        saveasjson(bigdinfo, self.reportfolder + 'bigdinfo.json')
        saveasjson(bigcontents, self.reportfolder + 'bigcontents.json')
        saveasjson(bigrelated, self.reportfolder + 'bigrelated.json')
        saveasjson(bigverlink, self.reportfolder + 'bigverlink.json')
        saveasjson(r_verlink, self.reportfolder + 'r_verlink.json')
        saveasjson(self.errors, self.reportfolder + 'bigerrors.json')

        with open(self.reportfolder + 'deletelist.txt', 'w', encoding='UTF-8') as fl:
            fl.write('\n'.join(sorted(set(self.deletelist))))

    def saveasXLS(self):
        colnames = ['ТОПИК', "ОШИБКА", "ИНФОРМАЦИЯ"]
        wb = opx.Workbook()
        wb.remove(wb.active)
        wb.add_named_style(tablehead)
        sheet = wb.create_sheet('ФАТАЛЬНЫЕ ОШИБКИ')
        r = 1
        c = 1
        for cn in colnames:
            sheet.cell(row=r, column=c, value=cn).style = tablehead
            column_letter = get_column_letter((c))
            sheet.column_dimensions[column_letter].width = 25
            c += 1
        for cellvalue in self.errors:
            if cellvalue == 'ROW':
                r += 1
                c = 1
            else:
                if cellvalue.find('\t') > -1:
                    cellvalue = cellvalue.split('\t')
                else:
                    cellvalue = [cellvalue]
                for cell in cellvalue:
                    val = re.sub(ILLEGAL_CHARACTERS_RE, '[?]', cell)
                    sheet.cell(row=r, column=c, value=val)
                    c += 1
        index = 0
        while True:
            if not os.path.exists(self.reportfolder + self.reportfile):
                break
            index += 1
            self.reportfile = os.path.splitext(self.reportfile)[0] + f' ({index}).xlsx'
        wb.save(self.reportfolder + self.reportfile)
        wb.close()
        print('Отчет сохранен: ' + self.reportfolder + self.reportfile)


class Killtr():
    def __init__(self):
        self.deletetopics = []
        self.files = []
        self.savedeleted = True
        self.savedir = ''
        self.processcount = cpu_count()
        self.filessize = 0

    def delete(self, topics):
        self.deletetopics = topics

    def addfiles(self, path):
        self.files.extend(glob.glob(path))

    def loadtopicsfromtxtfile(self, path):
        if not os.path.exists(path):
            print(f'не найден список топиков{path}')
            return
        with open(path, 'r', encoding='utf-8') as fl:
            lines = fl.readlines()
        result = []
        while True:
            if len(lines) == 0:
                break
            result.append(lines.pop(0).strip())
        if len(result) > 0:
            self.deletetopics = result
            print(f'найдено {len(result)} топиков для удаления')

    def gettopicsfromnsrfiles(self, mask):
        print(f'ищем топики для удаления в: {mask}')
        files = glob.glob(mask)
        result = []
        index = 0
        for file in files:
            index += 1
            print(index, len(files))
            with open(file, 'r', encoding='cp866', errors='ignore') as fl:
                lines = fl.readlines()
            for line in lines:
                if line != '' and line[0] == '!':
                    rsm = re.search('^!(TOPIC|OBJTOPIC|ANNOTOPIC)+\s+(\d+)', line)
                    if rsm:
                        result.append(rsm[2])
        self.deletetopics = result
        print(f'найдено топиков: {len(self.deletetopics)} шт.')

    def execute(self):
        procs = []
        filesqueue = Queue()
        if len(self.deletetopics) == 0:
            print('Список топиков для удаления пуст')
            return
        if len(self.files) == 0:
            print('Список нср файлов пуст')
            return
        for file in self.files:
            filesqueue.put(file)
            self.filessize += os.path.getsize(file)
        print(f'объем информации: {round(self.filessize / 1024 / 1024, 2)} m.b.')
        if self.savedeleted == True:
            print('режим сохранения удаленных топиков')
        else:
            print('топики удаляются без возможности восстановления')
        for i in range(0, self.processcount):
            proc = Process(target=killtr, args=(self.deletetopics, self.savedeleted, filesqueue,), daemon=True)
            # proc=Process(target=ki2, args=(f'Привет {i}',),daemon=True)
            proc.name = 'killer-' + str(i + 1)
            print(proc.name)
            procs.append(proc)
        for proc in procs:
            proc.start()

        # filesqueue.close()
        # filesqueue.join_thread()
        # salvqueue.join_thread()
        for proc in procs:
            print('join', proc.name)
            proc.join()
        print('killtr ok')


class Report():
    def __init__(self):
        print('init report')


class Excelreport():
    def __init__(self, name):
        self.wb = opx.Workbook()
        self.wb.remove(self.wb.active)
        self.sheets = self.wb.create_sheet(name)
        self.wb.add_named_style(tablehead)
        self.col = 1
        self.row = 1
        self.headers = []

    def addheaders(self, headers):
        self.headers = headers
        for header in headers:
            self.sheets.cell(row=self.row, column=self.col, value=header)
            if self.row == 2 and self.col == 1:
                self.sheets.cell(row=self.row, column=self.col, value=header).style = tablecolhead
            else:
                self.sheets.cell(row=self.row, column=self.col, value=header).style = tablehead
            column_letter = get_column_letter((self.col))
            self.sheets.column_dimensions[column_letter].width = 25
            self.col += 1

    def setcolwidth(self, widths):
        if len(self.headers) > 0:
            for col in range(0, len(widths)):
                column_letter = get_column_letter((self.col + 1))
                self.sheets.column_dimensions[column_letter].width = widths[col]

    def addrow(self, cells):
        self.col = 1
        self.row += 1
        for cell in cells:
            try:
                val = re.sub(ILLEGAL_CHARACTERS_RE, '[?]', cell)
                self.sheets.cell(row=self.row, column=self.col, value=val)
                self.col += 1
            except Exception as e:
                self.sheets.cell(row=self.row, column=self.col, value='error :(')
                self.col += 1

    def save(self, path):
        self.wb.save(path)
        self.wb.close()


class SearchInGkdumpiFdb():
    # класс в разработке
    def __init__(self, path2fdb, host, user, password):
        import fdb
        self.path2fdb = path2fdb
        self.host = host
        self.user = user
        self.password = password
        self.con = fdb.connect(dsn=host + ':' + self.path2fdb, user=user, password=password, charset='win1251')
        # self.con = fdb.connect(dsn='localhost:' + self.path2fdb, user='sysdba', password='i0r9b3i2s', charset='win1251')
        self.cursor = self.con.cursor()

    def showtables(self):
        # data = self.cursor.execute('select RDB$RELATION_NAME from RDB$RELATIONS where (RDB$SYSTEM_FLAG = 0)').fetchall()
        data = self.cursor.execute('''select R.RDB$RELATION_NAME, R.RDB$FIELD_POSITION, R.RDB$FIELD_NAME,
F.RDB$FIELD_LENGTH, F.RDB$FIELD_TYPE, F.RDB$FIELD_SCALE, F.RDB$FIELD_SUB_TYPE
from RDB$FIELDS F, RDB$RELATION_FIELDS R where F.RDB$FIELD_NAME = R.RDB$FIELD_SOURCE and R.RDB$SYSTEM_FLAG = 0
order by R.RDB$RELATION_NAME, R.RDB$FIELD_POSITION''').fetchall()
        print(data)
        for d in data:
            print(d)
        # data =
        # return self.cursor.fetchall()

    def test(self):
        sql = '''SELECT * FROM G_NAMES
        inner join g_status on g_names.GID = g_status.GID 
    inner join g_dates on g_names.gid=g_dates.GID'''
        self.cursor.execute(sql)

    def find_doc(self, parameters):
        params = []
        sql_opt = ''
        sql = 'select * FROM G_NAMES inner join G_STATUS on G_NAMES.GID = G_STATUS.GID'
        sql_opt += 'WHERE G_STATE=? '
        params.append('DS_DOC')

        dccd = parameters.get('code')
        if dccd:
            dccd = dccd.replace('З', '3').replace('О', '0').replace('I', '1')
            sql += ' inner join G_CODES on G_NAMES.GID = G_CODES.GID '
            sql_opt += ' AND G_CODE=? '
            params.append(dccd)
        if parameters.get('topic'):
            sql += ' AND G_NAMES.GID=? '
            params.append(parameters.get('topic'))
        date = parameters.get('date')
        if date:
            date = datetime.datetime.strptime(date, '%d.%m.%Y').date()
            print('date', type(date), date)
            sql += ' inner join G_DATES on G_NAMES.GID = G_DATES.GID '
            sql_opt += ' AND ddt=?'
            params.append(date)
        print(sql + ' ' + sql_opt)
        print(params)
        # return
        self.cursor.execute(sql + ' ' + sql_opt, params)
        data = self.cursor.fetchall()
        reliv = parameters.get('relevation')
        if reliv:
            for i, item in enumerate(data):
                item = list(item)
                percent = SM(None, reliv.upper(), item[3].upper()).ratio() * 100
                item.extend([str(percent), reliv])
                data[i] = tuple(item)
        return data


def parsensrc(sqlqueue, filesqueue, ):
    proc_name = current_process().name
    sqlqueue.put('add')
    while True:
        if filesqueue.qsize() == 0:
            break
        data = filesqueue.get()
        if data == 'break':
            filesqueue.put(data)
            break
        fl = open(data, 'r', encoding='cp866')
        lines = fl.readlines()
        fl.close()
        skip = False
        report = []
        topic = None
        dcd = 'Nocode'
        ddt = 'Nodate'
        dcnm = 'Noname'
        dmain = 'Nomain'
        related = 'Norelated'
        for line in lines:
            line = line.strip()
            if line == '':
                continue
            elif line in ['!EDITION', '!REL', '!NODOC']:
                skip = True
                topic = None
                dcd = 'Nocode'
                ddt = 'Nodate'
                dcnm = 'Noname'
                dmain = 'Nomain'
                related = 'Norelated'
            elif line[0] == '!':
                rsm = re.search('^!TOPIC (\d+)', line)
                if rsm:
                    topic = rsm[1]
                    dcd = 'Nocode'
                    ddt = 'Nodate'
                    dcnm = 'Noname'
                    dmain = 'Nomain'
                    related = 'Norelated'
                    skip = False
                if skip == True:
                    continue
                rsm = re.search('!RELATED (\d+)', line)
                if rsm:
                    related = rsm[1]
                rsm = re.search('!CODE (.+)', line)
                if rsm:
                    dcd = rsm[1].upper()
                rsm = re.search('!DATE (.+)', line)
                if rsm:
                    ddt = rsm[1]
                rsm = re.search('!MAIN (.+)', line)
                if rsm:
                    dmain = rsm[1]
                rsm = re.search('!NAME (.+)', line)
                if rsm:
                    if dcnm == 'Noname':
                        dcnm = rsm[1]
                    else:
                        dcnm = dcnm + ' ' + rsm[1]
                rsm = re.search('!STYLE', line)
                if rsm:
                    if skip == False and topic != None:
                        report.append([topic, ddt, dcd, dcnm, related, dmain])
                    skip = True
                    topic = None
                    dcd = 'Nocode'
                    ddt = 'Nodate'
                    dcnm = 'Noname'
                    dmain = 'Nomain'
                    related = 'Norelated'
            else:
                continue
        if skip == False:
            if topic:
                report.append([topic, ddt, dcd, dcnm, related, dmain])
        if len(report) > 0:
            sqlqueue.put(report)
            print(proc_name, 'осталось файлов:', filesqueue.qsize(), 'добавлено записей:', len(report))
    sqlqueue.put('remove')
    print(proc_name, 'завершен')


# функция вставляет данные в базу из очереди
def nsrc2sqlite(sqlqueue, dbpath):
    folder, file = os.path.split(dbpath)
    index = 0
    while os.path.exists(dbpath):
        try:
            os.remove(dbpath)
        except:
            index += 1
            newfile = str(index) + '-' + file
            dbpath = os.path.join(folder, newfile)

    conn = sqlite3.connect(dbpath)  # или :memory: чтобы сохранить в RAM
    cursor = conn.cursor()
    # Создание таблицы
    # topic,dcd,ddt,dcnm,related
    cursor.execute('PRAGMA synchronous = OFF')
    cursor.execute('PRAGMA journal_mode = OFF')
    cursor.execute("""CREATE TABLE CONTENTS
                      (topic text, ddt text, dcd text, related text,
                       dcnm text, dmain text)""")

    proc_name = current_process().name
    opencount = 0
    closecount = 0
    topiccount = 0
    while True:
        if opencount > 0 and closecount == opencount and sqlqueue.qsize() == 0:
            break
        if sqlqueue.qsize == 0:
            continue
        data = sqlqueue.get()
        if data == 'break':
            break
        elif data == 'add':
            opencount = opencount + 1
        elif data == 'remove':
            closecount = closecount + 1
        else:
            topiccount = topiccount + len(data)
            print(proc_name, topiccount)
            cursor.executemany('INSERT INTO contents VALUES (?,?,?,?,?,?)', data)
    print(proc_name, 'завершено')
    conn.commit()
    conn.close()


class SqLiteContents():
    def __init__(self, **kwargs):
        self.dbpath = kwargs.get('dbpath', os.getcwd() + 'contents.sqlite3')
        self.files = []

    def add_files(self, mask):
        temp = glob.glob(mask)
        self.files.extend(temp)
        print(f'Добавлено: {len(temp)} файлов')

    def create_contents(self):
        dt1 = datetime.datetime.now()

        sqlqueue = Queue()
        filesqueue = Queue()
        for file in self.files:
            filesqueue.put(file)
        print('в файловую очередь добавлены файлы:', filesqueue.qsize())
        self.procs = []
        cpucount = cpu_count()
        if cpucount > 1:
            cpucount -= 1
        for p in range(0, cpu_count() - 1):
            proc = Process(target=parsensrc, args=(sqlqueue, filesqueue,), daemon=False)
            proc.name = 'nsrcdump-' + str(p)
            proc.start()
            self.procs.append(proc)

            print('create:', proc.name)
        proc = Process(target=nsrc2sqlite, args=(sqlqueue, self.dbpath,), daemon=False)
        proc.name = 'sqlmanager'
        proc.start()
        time.sleep(2)
        for p in self.procs:
            print('join', p.name)
            p.join()
        print('завершено. очереди удалены')
        dt2 = datetime.datetime.now()
        delta = dt2 - dt1
        print('обновление нсрц заняло:', delta.seconds)
        self.createindex()

    # функция создает индексы внутри базы
    def createindex(self, ):
        print('create index...')
        conn = sqlite3.connect(self.dbpath)  # или :memory: чтобы сохранить в RAM
        cursor = conn.cursor()

        dt1 = datetime.datetime.now()
        cursor.execute('CREATE INDEX IF NOT EXISTS ddtidx ON CONTENTS (ddt)')
        dt2 = datetime.datetime.now()
        delta = dt2 - dt1
        print('индексирование дат заняло:', delta.seconds)

        cursor.execute('CREATE INDEX IF NOT EXISTS dcdidx ON CONTENTS (dcd)')
        dt3 = datetime.datetime.now()
        delta = dt3 - dt2
        print('индексирование номеров заняло:', delta.seconds)

        cursor.execute('CREATE INDEX IF NOT EXISTS gididx ON CONTENTS (topic)')
        dt4 = datetime.datetime.now()
        delta = dt4 - dt2
        print('индексирование топиков заняло:', delta.seconds)

        conn.commit()
        conn.close()
        print('транзакция завершена')

    def find_doc(self, parameters):
        def get_flag(params):
            if len(params) == 0:
                return ' WHERE '
            else:
                return ' AND '

        conn = sqlite3.connect(self.dbpath)  # или :memory: чтобы сохранить в RAM
        cursor = conn.cursor()
        sql = "SELECT * FROM CONTENTS"
        params = []
        if parameters.get('topic'):
            sql += f' {get_flag(params)} topic=?'
            params.append(parameters.get('topic'))
        if parameters.get('date'):
            sql += f' {get_flag(params)} ddt=?'
            params.append(parameters.get('date'))
        if parameters.get('code'):
            sql += f' {get_flag(params)} dcd=?'
            params.append(parameters.get('code').upper())
        if parameters.get('dmain'):
            sql += f' {get_flag(params)} dmain=?'
            params.append(parameters.get('dmain').upper())
        cursor.execute(sql, params)
        data = cursor.fetchall()
        reliv = parameters.get('relevation')
        if reliv:
            max_relive = 0
            max_item = None
            for i, item in enumerate(data):
                item = list(item)
                g_name = item[3]
                if parameters.get('crop_garant_name'):
                    g_name = crop_garant_name(g_name)
                percent = SM(None, reliv.upper(), g_name.upper()).ratio() * 100
                item.extend([percent, reliv])
                if  percent >  max_relive:
                    max_relive = percent
                    max_item = tuple(item)
                data[i] = tuple(item)
            # data = [max_item]
            if len(data) > 0:
                return [max_item,]

        if len(data) > 0:
            return data


if __name__ == '__main__':
    print('Этот модуль не является самостоятельной программой')
    exit()
